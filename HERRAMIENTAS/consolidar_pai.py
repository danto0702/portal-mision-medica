"""
Consolidador Registros Diarios de Vacunación — NIÑOS Y NIÑAS
ESE Hospital Regional Noroccidental · PAI 2025-2026
Agrega columnas MUNICIPIO y FECHA_VACUNACION
"""
import os, sys, re, datetime
import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Forzar UTF-8 en stdout para evitar errores de encoding en Windows
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE_DIR  = r'C:\Users\danto\Desktop\DANILO 2026\PAI\REGISTROS DIARIOS DE VACUNACIÓN\REGISTROS DIARIOS DE VACUNACIÓN'
OUTPUT    = r'C:\Users\danto\Desktop\DANILO 2026\PAI\Consolidado_NinosNinas_PAI.xlsx'
NCOLS     = 150
HDR_ROWS  = 10   # filas 1-10 son encabezado; datos desde fila 11

MESES = {
    'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
    'julio':7,'agosto':8,'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12
}

# Normalización sin acentos para comparar nombres de carpeta
def _norm(s):
    for a, b in [('Á','A'),('É','E'),('Í','I'),('Ó','O'),('Ú','U'),('Ñ','N'),
                 ('á','a'),('é','e'),('í','i'),('ó','o'),('ú','u'),('ñ','n')]:
        s = s.replace(a, b)
    return s.upper()

MUNI_KEYS = {
    'ABREGO': 'Ábrego',
    'CONVENCION': 'Convención',
    'EL CARMEN': 'El Carmen',
    'TEORAMA': 'Teorama',
}

def get_municipio(folder_name):
    n = _norm(folder_name)
    for key, val in MUNI_KEYS.items():
        if key in n:
            return val
    return folder_name  # devuelve el nombre original si no coincide

# ─── LECTURA ────────────────────────────────────────────
def find_sheet(names):
    """Busca la hoja NIÑOS Y NIÑAS tolerando variaciones de codificación."""
    for name in names:
        n = _norm(name)
        if 'NI' in n and 'AS' in n and 'Y' in n:
            return name
    return None

def read_file(filepath):
    ext = filepath.lower().rsplit('.', 1)[-1]
    try:
        if ext == 'xls':
            wb = xlrd.open_workbook(filepath)
            sname = find_sheet(wb.sheet_names())
            if not sname:
                return None
            sh = wb.sheet_by_name(sname)
            rows = []
            for ri in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    cell = sh.cell(ri, c)
                    # ctype 3 = XL_CELL_DATE → convertir a objeto date de Python
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode).date())
                        except:
                            row.append(cell.value)
                    else:
                        row.append(cell.value)
                rows.append(row)
            return rows
        else:  # xlsx
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sname = find_sheet(wb.sheetnames)
            if not sname:
                wb.close()
                return None
            rows = [list(row) for row in wb[sname].iter_rows(values_only=True)]
            wb.close()
            return rows
    except Exception as e:
        print(f'    ERROR al leer: {e}')
        return None

# ─── EXTRACCIÓN DE MES/AÑO ─────────────────────────────
def get_month_year(rows):
    # Busca en las filas 1-5 (índices 0-4); escanea TODAS las columnas
    for ri in range(min(6, len(rows))):
        row = rows[ri]
        for ci in range(len(row)):
            val = str(row[ci]).strip().lower()
            if val in MESES:
                month = MESES[val]
                # Busca el año en columnas adyacentes
                for delta in [1, 2, -1, 3]:
                    yi = ci + delta
                    if 0 <= yi < len(row):
                        try:
                            year = int(float(str(row[yi])))
                            if 2020 <= year <= 2030:
                                return month, year
                        except:
                            pass
    return None, None

# ─── IDENTIFICACIÓN DE FILA REAL ───────────────────────
def is_real_row(row):
    if len(row) < 4:
        return False
    # Col 0: CONSECUTIVO numérico >= 1
    try:
        if float(str(row[0])) < 1:
            return False
    except:
        return False
    # Col 3: identificación del padre/madre (no vacía)
    id_val = str(row[3]).strip()
    return id_val not in ('', '0', '0.0', 'nan', 'None')

# ─── NOMBRES DE COLUMNAS ───────────────────────────────
def build_col_names(rows):
    names = []
    # Combina contenido de filas 6, 7, 8 y 9 (índices 5-8)
    for col in range(NCOLS):
        parts = []
        for ri in [5, 6, 7, 8]:
            if ri < len(rows) and col < len(rows[ri]):
                v = str(rows[ri][col]).strip().replace('\n', ' ')
                v = re.sub(r'\s+', ' ', v)
                if v and v not in ('0', '0.0', 'nan', 'None') and v not in parts:
                    parts.append(v)
        names.append(' | '.join(parts) if parts else f'COL_{col+1}')
    return names

# ─── FECHA ─────────────────────────────────────────────
def make_date(day_val, month, year):
    # Ya es un objeto date (xlrd lo convirtió desde serial XL_CELL_DATE)
    if isinstance(day_val, datetime.date):
        return day_val
    if isinstance(day_val, datetime.datetime):
        return day_val.date()

    # String: puede ser "DD/MM/AAAA", "D/M/AAAA" o un número
    if isinstance(day_val, str):
        s = day_val.strip()
        for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y'):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except:
                pass
        try:
            day_val = float(s)   # puede ser "5" como string
        except:
            return None

    # Numérico
    try:
        fval = float(day_val)
        if 1 <= fval <= 31 and month and year:
            # Es el número del día del mes
            return datetime.date(year, month, int(fval))
        elif fval > 1000:
            # Serial de Excel que no fue convertido por xlrd (xlsx con data_only)
            import xlrd as _xlrd
            return _xlrd.xldate_as_datetime(fval, 0).date()
    except:
        pass
    return None

# ═══════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════
all_data   = []
col_names  = None
files_ok   = 0
files_skip = 0
resumen    = []  # para imprimir al final

print('=' * 70)
print('Consolidador PAI — NIÑOS Y NIÑAS  |  2025-2026')
print('=' * 70)

for muni_folder in sorted(os.listdir(BASE_DIR)):
    muni_path = os.path.join(BASE_DIR, muni_folder)
    if not os.path.isdir(muni_path):
        continue
    muni_name = get_municipio(muni_folder)
    muni_count = 0

    for year_folder in sorted(os.listdir(muni_path)):
        year_path = os.path.join(muni_path, year_folder)
        if not os.path.isdir(year_path):
            continue

        for fname in sorted(os.listdir(year_path)):
            if fname.startswith('~$'):
                continue
            ext = fname.lower().rsplit('.', 1)[-1]
            if ext not in ('xls', 'xlsx'):
                continue

            fpath = os.path.join(year_path, fname)
            print(f'  [{muni_name}] {year_folder}/{fname[:60]}')

            rows = read_file(fpath)
            if rows is None:
                print('    OMITIDO: hoja NIÑOS Y NIÑAS no encontrada')
                files_skip += 1
                continue

            month, year_num = get_month_year(rows)
            if not month or not year_num:
                print('    OMITIDO: no se pudo leer mes/año del encabezado')
                files_skip += 1
                continue

            if col_names is None:
                col_names = build_col_names(rows)

            count = 0
            for ri in range(HDR_ROWS, len(rows)):
                row = rows[ri]
                if is_real_row(row):
                    day    = row[1] if len(row) > 1 else None
                    fecha  = make_date(day, month, year_num)
                    padded = [(row[c] if c < len(row) else '') for c in range(NCOLS)]
                    all_data.append([muni_name, fecha] + padded)
                    count += 1

            print(f'    {count} registros — {month:02d}/{year_num}')
            muni_count += count
            files_ok += 1

    resumen.append(f'  {muni_name}: {muni_count} registros')

print()
print('─' * 70)
print('RESUMEN POR MUNICIPIO:')
for r in resumen:
    print(r)
print(f'TOTAL: {len(all_data)} registros de {files_ok} archivos ({files_skip} omitidos)')
print('─' * 70)

if not all_data:
    print('Sin datos. Verificar rutas y estructura de archivos.')
    sys.exit(1)

# ─── ESCRITURA DEL EXCEL ───────────────────────────────
print(f'\nEscribiendo archivo de salida...')

wb_out = Workbook(write_only=False)
ws = wb_out.active
ws.title = 'NIÑOS Y NIÑAS'

final_headers = ['MUNICIPIO', 'FECHA_VACUNACION'] + (col_names or [f'COL_{i}' for i in range(1, NCOLS+1)])

# Encabezado con estilo institucional
hdr_font  = Font(bold=True, color='FFFFFF', name='Arial', size=9)
hdr_fill  = PatternFill('solid', start_color='7c1d6f')
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.append(final_headers)
for c in range(1, len(final_headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font  = hdr_font
    cell.fill  = hdr_fill
    cell.alignment = hdr_align

ws.row_dimensions[1].height = 50

# Datos
date_fmt = 'DD/MM/YYYY'
data_font = Font(name='Arial', size=8)

for row_num, row_data in enumerate(all_data, start=2):
    ws.append(row_data)
    ws.cell(row=row_num, column=2).number_format = date_fmt

# Anchos de columna
ws.column_dimensions['A'].width = 14   # MUNICIPIO
ws.column_dimensions['B'].width = 17   # FECHA

# Inmovilizar paneles: desde columna C, fila 2
ws.freeze_panes = 'C2'

# Auto filtro en toda la tabla
ws.auto_filter.ref = f'A1:{get_column_letter(len(final_headers))}1'

wb_out.save(OUTPUT)
print(f'Archivo guardado: {OUTPUT}')
print('PROCESO COMPLETADO.')
